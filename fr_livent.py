# Copyright (C) 2020 coneypo
# SPDX-License-Identifier: MIT

# Author:   coneypo
# Blog:     http://www.cnblogs.com/AdaminXie
# GitHub:   https://github.com/coneypo/Dlib_face_recognition_from_camera
# Mail:     coneypo@foxmail.com

# 利用 OT 对于单张人脸追踪, 实时人脸识别 (Real-time face detection and recognition via Object-tracking for single face)

import dlib
import numpy as np
import cv2
import os
import pandas as pd
import time
from PIL import Image, ImageDraw, ImageFont
from openpyxl import Workbook
import time
import tkinter as tk
#threading引用必要套件
from threading import Thread
import threading

import openpyxl

import firebase_admin
from firebase_admin import credentials
from firebase_admin import firestore

import BotSpeak

# import serial
import time

import sys,os,dlib,glob,numpy
os.environ['TF_CPP_MIN_LOG_LEVEL'] = '2'
from skimage import io
import imutils
import pickle
from tensorflow.keras.preprocessing.image import img_to_array
from tensorflow.keras.models import load_model

import requests
# Dlib 正向人脸检测器 / Use frontal face detector of Dlib
detector = dlib.get_frontal_face_detector()

# Dlib 人脸 landmark 特征点检测器 / Get face landmarks
predictor = dlib.shape_predictor('data/data_dlib/shape_predictor_68_face_landmarks.dat')

# Dlib Resnet 人脸识别模型，提取 128D 的特征矢量 / Use Dlib resnet50 model to get 128D face descriptor
face_reco_model = dlib.face_recognition_model_v1("data/data_dlib/dlib_face_recognition_resnet_model_v1.dat")
# imSize=230

class Face_Recognizer:
    def __init__(self):
        self.font = cv2.FONT_ITALIC

        # 统计 FPS / For FPS
        self.frame_time = 0
        self.frame_start_time = 0
        self.fps = 0

        # 统计帧数 / cnt for frame
        self.frame_cnt = 0

        # 用来存储所有录入人脸特征的数组 / Save the features of faces in the database
        self.features_known_list = []
        # 用来存储录入人脸名字 / Save the name of faces in the database
        self.name_known_list = []

        # 用来存储上一帧和当前帧 ROI 的质心坐标 / List to save centroid positions of ROI in frame N-1 and N
        self.last_frame_centroid_list = []
        self.current_frame_centroid_list = []

        # 用来存储当前帧检测出目标的名字 / List to save names of objects in current frame
        self.current_frame_name_list = []

        # 上一帧和当前帧中人脸数的计数器 / cnt for faces in frame N-1 and N
        self.last_frame_faces_cnt = 0
        self.current_frame_face_cnt = 0

        # 用来存放进行识别时候对比的欧氏距离 / Save the e-distance for faceX when recognizing
        self.current_frame_face_X_e_distance_list = []

        # 存储当前摄像头中捕获到的所有人脸的坐标名字 / Save the positions and names of current faces captured
        self.current_frame_face_position_list = []
        # 存储当前摄像头中捕获到的人脸特征 / Save the features of people in current frame
        self.current_frame_face_feature_list = []

        # 控制再识别的后续帧数 / Reclassify after 'reclassify_interval' frames
        # 如果识别出 "unknown" 的脸, 将在 reclassify_interval_cnt 计数到 reclassify_interval 后, 对于人脸进行重新识别
        self.reclassify_interval_cnt = 0
        self.reclassify_interval = 10
        self.cap=cv2.VideoCapture(0)
        
        self.cap.set(3, 640) # set video widht
        self.cap.set(4, 480) # set video height
        self.close_flag=True
        #紀錄人臉框框座標
        self.img_top=0
        self.img_bottom=0
        self.img_right=0
        self.img_left=0
        self.color=(0,0,255)
        # self.bbb=[]
        # f = open('Class_Number.txt','r')
        # k = f.readlines()
        # f.close()
        # a=k[0].split(',')

        # for i in range(len(a)):
        #     self.bbb.append(a[i])
        self.IP = "192.168.11.102"
        fn = 'EE3407301.xlsx'
        wb = openpyxl.load_workbook(fn)
        wb.active = 0
        self.ws = wb.active
        self.Visitor_name=''
        self.find_who=''
        self.time=''
        self.place=''
        self.ID_bef=''
        self.Visitor_name_bef=''
        self.current_frame_face_cnt_bef=0
        self.a_bef=0
        self.bbb=0
        self.changeto=''
        cred = credentials.Certificate("./serviceAccount.json")
        firebase_admin.initialize_app(cred)

        self.db = firestore.client()

        self.ddd=100

        self.model = load_model("./livenet/liveness.model")
        self.le = pickle.loads(open("./livenet/le.pickle", "rb").read())
        self.net = cv2.dnn.readNetFromCaffe("./livenet/detector/deploy.prototxt.txt", "./livenet/detector/res10_300x300_ssd_iter_140000.caffemodel")
        self.label = ""
        self.pred = ""
        self.observed_resual="Unknow"
        self.correct_count=10
        self.REAL_THRESHOLD = 0.8 #will return fake if pred of real doesnt exceed threshold
        self.std_correct_time=0
        self.std_false_time=0
        self.observed_resual_singal=""
        self.observed_resual_bef= ""


        # COM_PORT = 'COM6'  # 請自行修改序列埠名稱
        # BAUD_RATES = 115200
        # self.ser = serial.Serial(COM_PORT, BAUD_RATES)

    def getLiveLabelfromImgandCoords(self,img, startX, startY, endX, endY, cw, ch):
        # global model,le,net,label,pred
        # lsy = startY
        # lsx = startX
        # ley = endY
        # lex = endX
        # fw = lex - lsx
        # fh = ley - lsy
        # rw = 1.3
        # rh = 0
        # (osx, osy, oex, oey) = (startX, startY, endX, endY)
        # if lsx - rw*fw > 0:
        #     lsx = int(lsx - rw*fw)
        # else:
        #     lsx = 0
        # if lsy - rh*fh > 0:
        #     lsy = int(lsy - rh*fh)
        # else:
        #     lsy = 0
        # if lex + rw*fw < cw:
        #     lex = int(lex + rw*fw)
        # else:
        #     lex = cw
        # if ley + rh*fh < ch:
        #     ley = int(ley + rh*fh)
        # else:
        #     ley = ch
        # liveFace = img[lsy:ley, lsx:lex]
        # liveFace = cv2.resize(liveFace, (32, 32))

        # liveFace = liveFace.astype("float") / 255.0
        # liveFace = img_to_array(liveFace)
        # liveFace = np.expand_dims(liveFace, axis=0)
        # preds = self.model.predict(liveFace)[0]

        # j = np.argmax(preds)

        # # self.label = self.le.classes_[j]
        # self.pred = str(round(preds[j],2))
        # cv2.rectangle(img, (lsx, lsy), (lex, ley),
		# 			(255, 255, 0), 2)
        # # cv2.rectangle(img, (lsx, lsy), (endX, endY),
        # #     (255, 255, 0), 2)
        lsy = startY
        lsx = startX
        ley = endY
        lex = endX
        fw = lex - lsx
        fh = ley - lsy
        rw = 2
        rh = 2
        if lsx - rw*fw > 0:
            lsx = int(lsx - rw*fw)
        else:
            lsx = 0
        if lsy - rh*fh > 0:
            lsy = int(lsy - rh*fh)
        else:
            lsy = 0
        if lex + rw*fw < cw:
            lex = int(lex + rw*fw)
        else:
            lex = cw
        if ley + rh*fh < ch:
            ley = int(ley + rh*fh)
        else:
            ley = ch
        liveFace = img[lsy:ley, lsx:lex]
        liveFace = cv2.resize(liveFace, (32, 32))

        liveFace = liveFace.astype("float") / 255.0
        liveFace = img_to_array(liveFace)
        liveFace = np.expand_dims(liveFace, axis=0)
        preds = self.model.predict(liveFace)[0]

        j = np.argmax(preds)

        self.label = self.le.classes_[j]
        self.pred = str(round(preds[j],2))

        if self.le.classes_[j] == "real" :
            # print(preds[j])
            if preds[j] > self.REAL_THRESHOLD:
                self.label = "real"
            else:
                self.label = "false"

        return self.label

    def bot_speak(self,sayword):
        BotSpeak.speak(sayword)
        if sayword == "請勿靠近閘門":
            sleep(10)
            BotSpeak.speak(sayword)
            # self.arduino_server(0)

    # 从 "features_all.csv" 读取录入人脸特征 / Get known faces from "features_all.csv"
    def get_face_database(self):
        if os.path.exists("data/features_all.csv"):
            path_features_known_csv = "data/features_all.csv"
            csv_rd = pd.read_csv(path_features_known_csv, header=None)
            for i in range(csv_rd.shape[0]):
                features_someone_arr = []
                for j in range(0, 128):
                    if csv_rd.iloc[i][j] == '':
                        features_someone_arr.append('0')
                    else:
                        features_someone_arr.append(csv_rd.iloc[i][j])
                self.features_known_list.append(features_someone_arr)
                self.name_known_list.append("Person_" + str(i + 1))
            #print("Faces in Database：", len(self.features_known_list))
            return 1
        else:
            print('##### Warning #####', '\n')
            print("'features_all.csv' not found!")
            print(
                "Please run 'get_faces_from_camera.py' and 'features_extraction_to_csv.py' before 'face_reco_from_camera.py'",
                '\n')
            print('##### End Warning #####')
            return 0

    # 更新 FPS / Update FPS of video stream
    def update_fps(self):
        now = time.time()
        self.frame_time = now - self.frame_start_time
        self.fps = 1.0 / self.frame_time
        self.frame_start_time = now

    # 计算两个128D向量间的欧式距离 / Compute the e-distance between two 128D features
    @staticmethod
    def return_euclidean_distance(feature_1, feature_2):
        feature_1 = np.array(feature_1)
        feature_2 = np.array(feature_2)
        dist = np.sqrt(np.sum(np.square(feature_1 - feature_2)))
        return dist

    # 生成的 cv2 window 上面添加说明文字 / putText on cv2 window
    '''def draw_note(self, img_rd):
        # 添加说明 (Add some statements
        #cv2.putText(img_rd, "Face Recognizer with OT (one person)", (20, 40), self.font, 1, (255, 255, 255), 1, cv2.LINE_AA)
        cv2.putText(img_rd, "FPS:   " + str(self.fps.__round__(2)), (20, 100), self.font, 0.8, (0, 255, 0), 1,
                    cv2.LINE_AA)
        cv2.putText(img_rd, "Q: Quit", (20, 450), self.font, 0.8, (255, 255, 255), 1, cv2.LINE_AA)'''
    def change_firebase(self,ID,room):
        if(self.ID_bef != ID and ID != 'unknown' ):
            self.ID_bef=ID
            # print("change：{:}   {:}".format(ID,room))
            doc_ref = self.db.collection('visit people').document('visit people')
            docs = doc_ref.get()
            doc_bef=docs.to_dict()
            ID_i='{:}'.format(ID)
            Room='{:}'.format(room)
            doc_bef[ID_i]=Room
            doc_ref_visit = self.db.collection("visit people").document("visit people")
            doc_ref_visit.set(doc_bef)
    def arduino_close(self):
        # print(time.ctime())
        time.sleep(10)
        # print(time.ctime())
        print("close....")
        # self.bot_speak("門即將關閉")
        # Send = "http://{:}/OFF".format(self.IP)
        # Thread(target=self.semd_line_msg,args =(Send,)).start()
        # Thread(target=self.semd_line_msg,args =(Send,)).start()

    def arduino_server(self,a):
        # print(a)
        if(self.a_bef != a):
            self.a_bef = a
            # print(a)
            if(a==90):
                print("open....")
                # Send = "http://{:}/ON".format(self.IP)
                # Thread(target=self.semd_line_msg,args =(Send,)).start()
                # Thread(target=self.semd_line_msg,args =(Send,)).start()
                # self.ser.write(b'open\n')
            else:
                Thread(target=self.arduino_close,).start() 
                # self.ser.write(b'close\n')
    def semd_line_msg(self,send):
        r = requests.get(send) 

    def draw_name(self, img_rd):
        # 在人脸框下面写人脸名字 / Write names under ROI
        #print(self.current_frame_name_list)
        font = ImageFont.truetype("simsun.ttc", 30)
        '''img = Image.fromarray(cv2.cvtColor(img_rd, cv2.COLOR_BGR2RGB))
        draw = ImageDraw.Draw(img)
        
        draw.text(xy=self.current_frame_face_position_list[0], text=aaa[1], font=font)
        img_rd = cv2.cvtColor(np.array(img), cv2.COLOR_RGB2BGR)'''
        


        aaa=self.current_frame_name_list[0].split("_")
        if len(aaa)!=2:aaa=["","0"]
        if int(aaa[1])+1 != 1:
            read_Visitor_name='A'+str(int(aaa[1])+1)
            read_find_who='D'+str(int(aaa[1])+1)
            read_time='E'+str(int(aaa[1])+1)
            read_place='F'+str(int(aaa[1])+1)
            self.Visitor_name = self.ws[read_Visitor_name].value
            self.find_who = self.ws[read_find_who].value
            self.time = self.ws[read_time].value
            self.place = self.ws[read_place].value
            # self.color = (0,255,0)
        else:
            self.Visitor_name = 'unknown'  
            self.find_who = 'unknown'  
            self.time = 'unknown'  
            self.place = 'unknown'  
            # self.color = (0,0,255)
        if(abs(self.img_top - self.img_bottom)>self.ddd):
            
            cv2.putText(img_rd, self.Visitor_name, self.current_frame_face_position_list[0], self.font, 0.8, self.color, 1, cv2.LINE_AA)
            # if(self.Visitor_name == 'unknown' or self.observed_resual == 'false'):self.arduino_server(0)
            # if(self.Visitor_name != 'unknown' and self.observed_resual == 'real'):self.arduino_server(90)
            if(self.Visitor_name != 'unknown'):
                if(self.Visitor_name_bef != self.Visitor_name):
                    
                    self.Visitor_name_bef = self.Visitor_name
                    #Get Doc
                    self.doc_ref = self.db.collection('meeting room').document('meeting room')
                    self.docs = self.doc_ref.get()
                    
                    ccc='{:}'.format(self.place)
                    self.bbb=self.docs.to_dict().get(ccc)
                    # print('{:}會議室 => {}'.format(self.place,bbb))
                    # if self.bbb == 0:
                    #     www="學號{:}，確認身分，請進入閘門，會議室代號為{:}".format(self.Visitor_name,self.place)
                    #     # www="學號{:}，確認身分，請進入閘門，會議室代號為{:}".format(self.Visitor_name,self.place)
                    # else:
                    #     if(self.docs.to_dict().get('A')==0):self.changeto='A'
                    #     elif(self.docs.to_dict().get('B')==0):self.changeto='B'
                    #     elif(self.docs.to_dict().get('C')==0):self.changeto='C'
                    #     # print("changeto = ",changeto)
                    #     self.change_firebase(self.Visitor_name,self.changeto)
                    #     # www="學號{:}，確認身分，請進入閘門，會議室代號更改為{:}".format(self.Visitor_name,self.changeto)
                    
                    
                # if self.bbb == 0:    
                #     cv2.putText(img_rd, "Meeting Room "+self.place , (20, 100), self.font, 0.8, (0, 255, 0), 1,
                #             cv2.LINE_AA)   
                # else:
                #     cv2.putText(img_rd, "Meeting Room "+self.place+" --> "+ self.changeto, (20, 100), self.font, 0.8, (0, 255, 0), 1,
                #             cv2.LINE_AA) 
                if(self.Visitor_name != 'unknown'):
                    # stX=int(self.img_left/2)
                    # stY=int(self.img_top/2)
                    # enX=int(self.img_left/2)
                    # enY=int(self.img_left/2)
                    
                    self.observed_resual_singal=self.getLiveLabelfromImgandCoords(img_rd,self.img_left,self.img_top,self.img_right,self.img_bottom,1024,768) 
                    print(self.observed_resual_singal)  
                else:
                    self.observed_resual_singal="UnKnow"
                
                if self.observed_resual_singal == "real":
                    self.std_correct_time+=1 
                    self.std_false_time=0
                else:  
                    self.std_false_time+=1                               
                    self.std_correct_time=0

                    # if self.observed_resual != 'false':
                # print("{:}  {:}".format(self.observed_resual,self.observed_resual_bef))    
                if self.observed_resual !=  self.observed_resual_bef:   

                    
                    print("in")
                    
                    if  self.std_correct_time>self.correct_count:
                        self.observed_resual_bef = self.observed_resual
                        self.observed_resual='real'
                        self.color = (0,255,0)
                        www="學號{:}，確認身分，請進入閘門，會議室代號為{:}".format(self.Visitor_name,self.place)    
                        Thread(target=self.bot_speak,args =(www,)).start()   
                        Send = "https://31a6b5649d69.ngrok.io/push_function/{:}您好：\n您有一名訪客來臨\n訪客性名：{:}\n抵達時間：{:}\n會議室地點：{:}會議室/Uf0ed3942c6fa32fba9b380cefd930491".format(self.find_who,self.Visitor_name,time.strftime("%Y-%m-%d %H:%M:%S", time.localtime()),self.place)
                        Thread(target=self.semd_line_msg,args =(Send,)).start()  
                        self.arduino_server(90)
                    # else:    
                    elif(self.std_false_time>self.correct_count):
                        self.observed_resual_bef = self.observed_resual
                        if self.observed_resual != 'false' and self.Visitor_name != 'unknown':
                            wwww ="學號{:}請勿試圖闖關".format(self.Visitor_name)
                            Thread(target=self.bot_speak,args =(wwww,)).start() 
                        self.observed_resual='false'
                        self.color = (0,0,255)    
                    # else:  
                    #     self.observed_resual='wait'
                    #     self.color = (0,255,255)
                    # if(self.observed_resual_singal !="real"):  
                    #     if()                
                    #     wwww ="學號{:}請勿試圖闖關".format(self.Visitor_name)
                    #     Thread(target=self.bot_speak,args =(wwww,)).start()     
                    # print("observed_resual = ",self.observed_resual)                          
                    # if(abs(self.img_top - self.img_bottom)<100):self.arduino_server(0)
                    # if(self.current_frame_face_cnt == 1):print("face....")
                    # else:print("no face....")
                    # print(abs(self.img_top - self.img_bottom))


        # print(self.place)
        #print("bbb[aaa[1]] = {:}".format(bbb[int(aaa[1])]))
        return img_rd

    def show_chinese_name(self):
        # Default known name: person_1, person_2, person_3
        if self.current_frame_face_cnt >= 1:
            #print(self.name_known_list)
            '''self.name_known_list[0] = '张1'.encode('utf-8').decode()
            self.name_known_list[1] = '张2'.encode('utf-8').decode()
            self.name_known_list[2] = '张3'.encode('utf-8').decode()
            self.name_known_list[3] = '张4'.encode('utf-8').decode()
            self.name_known_list[4] = '张5'.encode('utf-8').decode()'''
            self.name_known_list.append('张1'.encode('utf-8').decode())
            self.name_known_list.append('张2'.encode('utf-8').decode())
            self.name_known_list.append('张3'.encode('utf-8').decode())
            self.name_known_list.append('张4'.encode('utf-8').decode())
            self.name_known_list.append('张5'.encode('utf-8').decode())

    # 处理获取的视频流，进行人脸识别 / Face detection and recognition wit OT from input video stream
    def process(self, stream):
        # 1. 读取存放所有人脸特征的 csv / Get faces known from "features.all.csv"
        if self.get_face_database():
            while stream.isOpened() and self.close_flag :
                self.frame_cnt += 1
                #print(">>> Frame " + str(self.frame_cnt) + " starts")
                flag, img_rd = stream.read()
                kk = cv2.waitKey(1)

                # 2. 检测人脸 / Detect faces for frame X
                faces = detector(img_rd, 0)

                # 3. 更新帧中的人脸数 / Update cnt for faces in frames
                self.last_frame_faces_cnt = self.current_frame_face_cnt
                self.current_frame_face_cnt = len(faces)

                # 4.1 当前帧和上一帧相比没有发生人脸数变化 / If cnt not changes, 1->1 or 0->0
                if self.current_frame_face_cnt == self.last_frame_faces_cnt:
                    #print("   >>> scene 1: 当前帧和上一帧相比没有发生人脸数变化 / No face cnt changes in this frame!!!")
                    if "unknown" in self.current_frame_name_list:
                        #print("   >>> 有未知人臉, 開始進行 reclassify_interval_cnt 計數 {:}".format(self.reclassify_interval_cnt))
                        self.reclassify_interval_cnt += 1

                    # 4.1.1 当前帧一张人脸 / One face in this frame
                    if self.current_frame_face_cnt ==1:
                        if self.reclassify_interval_cnt==self.reclassify_interval:
                            #print("   >>> scene 1.1 需要对于当前帧重新进行人脸识别 / Re-classify for current frame")

                            self.reclassify_interval_cnt=0
                            self.current_frame_face_feature_list = []
                            self.current_frame_face_X_e_distance_list = []
                            self.current_frame_name_list = []

                            for i in range(len(faces)):
                                shape = predictor(img_rd, faces[i])
                                self.current_frame_face_feature_list.append(
                                    face_reco_model.compute_face_descriptor(img_rd, shape))

                            # a. 遍历捕获到的图像中所有的人脸 / Traversal all the faces in the database
                            for k in range(len(faces)):
                                self.current_frame_name_list.append("unknown")

                                # b. 每个捕获人脸的名字坐标 / Positions of faces captured
                                self.current_frame_face_position_list.append(tuple(
                                    [faces[k].left(),
                                     int(faces[k].bottom() + (faces[k].bottom() - faces[k].top()) / 4)]))

                                # c. 对于某张人脸，遍历所有存储的人脸特征 / For every face detected, compare it with all the faces in the database
                                for i in range(len(self.features_known_list)):
                                    # 如果 person_X 数据不为空 / If the data of person_X is not empty
                                    if str(self.features_known_list[i][0]) != '0.0':
                                        #print("            >>> with person", str(i + 1), "the e distance: ", end='')
                                        e_distance_tmp = self.return_euclidean_distance(
                                            self.current_frame_face_feature_list[k],
                                            self.features_known_list[i])
                                        #print(e_distance_tmp)
                                        self.current_frame_face_X_e_distance_list.append(e_distance_tmp)
                                    else:
                                        # 空数据 person_X / For empty data
                                        self.current_frame_face_X_e_distance_list.append(999999999)
                                #print("            >>> current_frame_face_X_e_distance_list:", self.current_frame_face_X_e_distance_list)

                                # d. 寻找出最小的欧式距离匹配 / Find the one with minimum e distance
                                similar_person_num = self.current_frame_face_X_e_distance_list.index(
                                    min(self.current_frame_face_X_e_distance_list))

                                if min(self.current_frame_face_X_e_distance_list) < 0.4:
                                    # 在这里更改显示的人名 / Modify name if needed
                                    self.show_chinese_name()
                                    self.current_frame_name_list[k] = self.name_known_list[similar_person_num]
                                    #print("            >>> recognition result for face " + str(k + 1) + ": " +self.name_known_list[similar_person_num])
                                #else:
                                    #print("            >>> recognition result for face " + str(k + 1) + ": " + "unknown")
                        else:
                            #print("   >>> scene 1.2 不需要对于当前帧重新进行人脸识别 / No re-classification for current frame")
                            # 获取特征框坐标 / Get ROI positions

                            for k, d in enumerate(faces):
                                # 计算矩形框大小 / Compute the shape of ROI
                                self.img_top=d.top()
                                self.img_bottom=d.bottom()
                                self.img_right=d.right()
                                self.img_left=d.left()
                                height = (d.bottom() - d.top())
                                width = (d.right() - d.left())
                                hh = int(height / 2)
                                ww = int(width / 2)
                                if(abs(self.img_top - self.img_bottom)>self.ddd):
                                    cv2.rectangle(img_rd,
                                                tuple([d.left() - ww, d.top() - hh]),
                                                tuple([d.right() + ww, d.bottom() + hh]),
                                                self.color, 2)

                                self.current_frame_face_position_list[k] = tuple(
                                    [faces[k].left(), int(faces[k].bottom() + (faces[k].bottom() - faces[k].top()) / 4)])

                                #print("   >>> self.current_frame_name_list:              ",self.current_frame_name_list[k])
                                #print("   >>> self.current_frame_face_position_list:     ",self.current_frame_face_position_list[k])

                                img_rd = self.draw_name(img_rd)
                # 4.2 当前帧和上一帧相比发生人脸数变化 / If face cnt changes, 1->0 or 0->1
                else:
                    #print("   >>> scene 2: 当前帧和上一帧相比人脸数发生变化 / Faces cnt changes in this frame")
                    self.current_frame_face_position_list = []
                    self.current_frame_face_X_e_distance_list = []
                    self.current_frame_face_feature_list = []

                    # 4.2.1 人脸数从 0->1 / Face cnt 0->1
                    if self.current_frame_face_cnt == 1:
                        self.current_frame_face_cnt_bef = self.current_frame_face_cnt
                        #print("   >>> scene 2.1 出现人脸，进行人脸识别 / Get person in this frame and do face recognition")
                        self.current_frame_name_list = []

                        for i in range(len(faces)):
                            shape = predictor(img_rd, faces[i])
                            self.current_frame_face_feature_list.append(
                                face_reco_model.compute_face_descriptor(img_rd, shape))

                        # a. 遍历捕获到的图像中所有的人脸 / Traversal all the faces in the database
                        for k in range(len(faces)):
                            self.current_frame_name_list.append("unknown")

                            # b. 每个捕获人脸的名字坐标 / Positions of faces captured
                            self.current_frame_face_position_list.append(tuple(
                                [faces[k].left(), int(faces[k].bottom() + (faces[k].bottom() - faces[k].top()) / 4)]))

                            # c. 对于某张人脸，遍历所有存储的人脸特征 / For every face detected, compare it with all the faces in database
                            for i in range(len(self.features_known_list)):
                                # 如果 person_X 数据不为空 / If data of person_X is not empty
                                if str(self.features_known_list[i][0]) != '0.0':
                                    #print("            >>> with person", str(i + 1), "the e distance: ", end='')
                                    e_distance_tmp = self.return_euclidean_distance(
                                        self.current_frame_face_feature_list[k],
                                        self.features_known_list[i])
                                    #print(e_distance_tmp)
                                    self.current_frame_face_X_e_distance_list.append(e_distance_tmp)
                                else:
                                    # 空数据 person_X / Empty data for person_X
                                    self.current_frame_face_X_e_distance_list.append(999999999)

                            # d. 寻找出最小的欧式距离匹配 / Find the one with minimum e distance
                            similar_person_num = self.current_frame_face_X_e_distance_list.index(min(self.current_frame_face_X_e_distance_list))

                            if min(self.current_frame_face_X_e_distance_list) < 0.4:
                                # 在这里更改显示的人名 / Modify name if needed
                                self.show_chinese_name()
                                self.current_frame_name_list[k] = self.name_known_list[similar_person_num]
                                #print("            >>> recognition result for face " + str(k + 1) + ": " +self.name_known_list[similar_person_num])
                            #else:
                                #print("            >>> recognition result for face " + str(k + 1) + ": " + "unknown")

                        if "unknown" in self.current_frame_name_list:
                            self.reclassify_interval_cnt+=1

                    # 4.2.1 人脸数从 1->0 / Face cnt 1->0
                    elif self.current_frame_face_cnt == 0:
                        #print("   >>> scene 2.2 人脸消失, 当前帧中没有人脸 / No face in this frame!!!")
                        # if(abs(self.img_top - self.img_bottom)>150):
                        if(self.current_frame_face_cnt_bef != self.current_frame_face_cnt):
                            self.current_frame_face_cnt_bef = self.current_frame_face_cnt
                            self.Visitor_name_bef = ''
                            self.observed_resual_bef = '1231'
                            self.color = (0,0,255)
                            self.arduino_server(0)
                            # print("clear")
                            # Thread(target=self.bot_speak,args =("門即將關閉",)).start()
                        self.reclassify_interval_cnt=0
                        self.current_frame_name_list = []
                        self.current_frame_face_feature_list = []

                # 5. 生成的窗口添加说明文字 / Add note on cv2 window
                #self.draw_note(img_rd)
                
                if kk == ord('q'):
                    break
    

                #self.update_fps()
                img_rd = cv2.resize(img_rd, (1024 ,768),interpolation=cv2.INTER_CUBIC)
                h1,w1,l = np.shape(img_rd)
                w=int((screenwidth-w1)/2)
                h=int((screenheight-h1)/2)
                cv2.namedWindow("camera", 1)
                # cv2.moveWindow('camera',w,h)
                cv2.imshow("camera", img_rd)
                

                #print(">>> Frame ends\n\n")

    def run(self):
        # self.cap = cv2.VideoCapture(0)
        self.process(self.cap)

        self.cap.release()
        cv2.destroyAllWindows()

def main(sw,sh):
    global screenwidth,screenheight
    screenwidth=sw
    screenheight=sh
    print(screenwidth,screenheight)
    Face_Recognizer_con = Face_Recognizer()
    Face_Recognizer_con.run()
def main1():
    global screenwidth,screenheight
    screenwidth=1920
    screenheight=1080
    print(screenwidth,screenheight)
    Face_Recognizer_con = Face_Recognizer()
    Face_Recognizer_con.run()    

if __name__ == '__main__':
    
    main1()
    