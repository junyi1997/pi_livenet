import tkinter as tk
from tkinter import ttk
import tkinter.messagebox as messagebox
import pickle
from PIL import Image,ImageTk
from tkinter import Scrollbar, Frame
from tkinter.ttk import Treeview
# import BotSpeak
# list_name = ["Adela Chen","Bonnie Lin","Allan Lin","M10907716","M10907306","M10907324"]
# list_visit = ["Junyi Wu","Baron Syu","Frank Zhou","M10907324","M10907324","M10907324"]
# list_time = ["2021/08/03/15","2021/08/04/15","2021/08/05/15","2021/08/06/15","2021/08/07/15","2021/08/08/15"]
import fr_livent
import openpyxl

########################################################################
class MyApp(object):
    """"""
    #----------------------------------------------------------------------
    def __init__(self):
        self.list_name=[]
        self.list_visit=[]
        self.list_time=[]
        """Constructor"""
        self.win = tk.Tk()
        # self.win.attributes("-fullscreen", True)
        self.win.geometry("1024x768")
        self.win.title("雲端註冊與深度辨識真面目系統")#定義標題名稱
        left=(self.win.winfo_screenwidth()-1024)//2#指定視窗置中
        top=(self.win.winfo_screenheight()-768)//2
        self.win.geometry("{:}x{:}+{:}+{:}".format(1024,768,left,top))

        print(self.win.winfo_screenwidth(),self.win.winfo_screenheight())
        #圖片呼叫
        self.photo_background=tk.PhotoImage(file=r"./image/main_background.png")
        self.photo_inSystem=tk.PhotoImage(file=r"./image/in_system.png")
        self.photo_inSearch=tk.PhotoImage(file=r"./image/in_search.png")

        canvas_width = 1024#新增一個畫布
        canvas_height =768
        canvas = tk.Canvas(self.win, 
        width=canvas_width, 
        height=canvas_height)
        canvas.pack()
        #背景
        canvas.create_image(512,384, image=self.photo_background)#將背景貼到畫布上
        
        def inSystem():
            # self.hide()
            fr_livent.main(self.win.winfo_screenwidth(),self.win.winfo_screenheight())
            # self.show()
            
        #選擇使用說明
        but_System=tk.Button(self.win,image=self.photo_inSystem, command=inSystem) 
        but_System.place(x=50,y=600)
        
        #進入登入畫面
        but_Search=tk.Button(self.win,image=self.photo_inSearch, command=self.openFrame) 
        but_Search.place(x=600,y=600)
        # BotSpeak.speak("歡迎來到KEEPING個人資料管理系統  請點選下方按鍵登入")
        self.win.mainloop()
        # BotSpeak.speak("掰掰") 

    #----------------------------------------------------------------------
    def hide(self):
        """"""
        self.win.withdraw()
    
    def closeWindow(self,myclosewindow):
        self.onCloseOtherFrame(myclosewindow)

    def getinfo(self):
        fn = 'EE3407301.xlsx'
        wb = openpyxl.load_workbook(fn)
        wb.active = 0
        ws = wb.active
        print(ws.max_row)
        print(ws.max_column)
        # print(wb)
        for i in range(int(ws.max_row-1)):
            read_Visitor_name='A'+str(i+2)
            read_find_who='D'+str(i+2)
            read_time='E'+str(i+2)
            read_place='F'+str(i+2)
            self.list_name.append(ws[read_Visitor_name].value)
            self.list_visit.append(ws[read_find_who].value)
            self.list_time.append(ws[read_time].value)
        # print("list_name = {:}".format(self.list_name))   
        # print("list_visit = {:}".format(self.list_visit))  
        # print("list_time = {:}".format(self.list_time))  
         

    # #----------------------------------------------------------------------
    def openFrame(self):

        """"""
        self.hide()
        self.win_Search = tk.Toplevel()
        # self.win_Search.attributes("-fullscreen", True)
        #使用者關閉視窗觸發的事件（第一個刪除視窗，第二個為函式名，即過程）
        self.win_Search.protocol('WM_DELETE_WINDOW',lambda:self.closeWindow(self.win_Search))
        #win_Search.attributes("-fullscreen", True)
        left=(self.win_Search.winfo_screenwidth()-1024)//2
        top=(self.win_Search.winfo_screenheight()-768)//2
        self.win_Search.geometry("{:}x{:}+{:}+{:}".format(1024,768,left,top))
        self.win_Search.title("行事曆")
        self.win_Search.photo_background=tk.PhotoImage(file=r"./image/Search_background.png")
        self.win_Search.photo_back=tk.PhotoImage(file=r"./image/back.png")

        canvas_width = 1024
        canvas_height =768
        canvas = tk.Canvas(self.win_Search, 
        width=canvas_width, 
        height=canvas_height)
        canvas.pack()
        #背景
        canvas.create_image(512,384, image=self.win_Search.photo_background)

        btn01= tk.Button(self.win_Search,image=self.win_Search.photo_back,command=lambda: self.onCloseOtherFrame(self.win_Search) )
        btn01.place(x=800,y=670)

                #使用Treeview組件實現表格功能

        frame = Frame(self.win_Search)

        frame.place(x=50, y=50, width=800, height=600)

        style_head = ttk.Style()
        style_head.configure("Treeview.Heading", font=('Noto Sans Mono CJK TC Bold', 30), rowheight=200)
        style_head.configure("Treeview", font=('Noto Sans Mono CJK TC Bold', 30), rowheight=100)

        #滾動條

        scrollBar = tk.Scrollbar(frame)

        scrollBar.pack(side=tk.RIGHT, fill=tk.Y)

        #Treeview組件，6列，顯示表頭，帶垂直滾動條

        tree = Treeview(frame,

                                columns=( 'c1' , 'c2' , 'c3' ),

                                show= "headings" ,

                                yscrollcommand=scrollBar.set)

        #設置每列寬度和對齊方式

        tree.column( 'c1' , width=230, anchor= 'center' )

        tree.column( 'c2' , width=230, anchor= 'center' )

        tree.column( 'c3' , width=340, anchor= 'center' )

        #設置每列表頭標題文本

        tree.heading( 'c1' , text= '訪客姓名' )

        tree.heading( 'c2' , text= '受訪者姓名' )

        tree.heading( 'c3' , text= '來訪時間' )

        tree.pack(side=tk.LEFT, fill=tk.Y)

        #Treeview組件與垂直滾動條結合

        scrollBar.config(command=tree.yview)

        #定義並綁定Treeview組件的鼠標單擊事件

        def treeviewClick(event):

            pass

        tree.bind( '<Button-1>' , treeviewClick)
        # print(len(list_time))
        self.getinfo()
        for i in range(len(self.list_name)):
            tree.insert("",i,values=(self.list_name[i],self.list_visit[i],self.list_time[i])) #插入數據
        # tree.insert("",1,values=("Adela Chen","Junyi Wu","2021/08/03/15")) #插入數據
        # tree.insert("",2,values=("Bonnie Lin","Baron Syu","2021/08/04/15")) #插入數據
        # tree.insert("",3,values=("Allan Lin","Frank Zhou","2021/08/05/15")) #插入數據

    #----------------------------------------------------------------------
    def onCloseOtherFrame(self, otherFrame):
        """"""
        otherFrame.destroy()
        self.show()
    #----------------------------------------------------------------------
    def CloseWin(self, otherFrame):
        """"""
        otherFrame.destroy()

    #----------------------------------------------------------------------
    def show(self):
        """"""
        self.win.update()
        self.win.deiconify()
#----------------------------------------------------------------------
if __name__ == "__main__":
    app = MyApp() 